from openpyxl import Workbook  # 「pip install openpyxl」でインストールしておく
import streamlit as st
import cv2
from pyzbar.pyzbar import decode
import numpy as np

# あらかじめ登録されているバーコード情報
barcode_database = {
    "0000000001": "秋岡義幸",
    # ここにさらにバーコードと商品名を追加できます
}

st.title('バーコード読み取りアプリ')

# 画像アップロードを受け付ける
uploaded_file = st.file_uploader("画像をアップロードしてください", type=['png', 'jpg', 'jpeg'])

def read_barcode(image):
    decoded_objects = decode(image)
    for obj in decoded_objects:
        return obj.data.decode("utf-8"), obj.rect
    return None, None
i = ""
if uploaded_file is not None:
    file_bytes = np.asarray(bytearray(uploaded_file.read()), dtype=np.uint8)
    image = cv2.imdecode(file_bytes, 1)
    
    # バーコードのデータを読み取る
    barcode_data, rect = read_barcode(image)

    if barcode_data:
        # 読み取ったバーコード情報を枠で囲む
        if rect:
            cv2.rectangle(image, (rect.left, rect.top), (rect.left + rect.width, rect.top + rect.height), (255, 0, 0), 2)
        st.image(image, channels="BGR")

        if barcode_data in barcode_database:
            product_name = barcode_database[barcode_data]
            i = barcode_data
            edited_data = st.text_input("商品の名前を編集", value=product_name)
            
            # データの保存
            if st.button("データを保存"):
                barcode_database[barcode_data] = edited_data
                st.write("保存されたデータ:", edited_data)
                st.success("データが保存されました。")
        else:
            st.write("このバーコードはデータベースに登録されていません。")
    else:
        st.image(image, channels="BGR")
        st.write("バーコードが読み取れませんでした。")

# ワークブックの新規作成と保存
wb = Workbook()
wb.save('myworkbook.xlsx')

# ワークブックの読み込み
from openpyxl import load_workbook

wb = load_workbook('バーコード.xlsx')

# ワークシートの選択
ws = wb['Sheet']  # ワークシートを指定
ws = wb.active  # アクティブなワークシートを選択
print(f'sheet name: {ws.title}')  # sheet name: Sheet
ws['A' + i ] = 'Hello from Python'
# ワークシートの作成
wb.create_sheet('my sheet')


wb.save('バーコード.xlsx')
